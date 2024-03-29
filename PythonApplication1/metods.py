import matplotlib
import matplotlib as mpl 
import matplotlib.pyplot as plt 
import math, random
import pandas as pd
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import tensorflow
import tensorflow as tf
from tensorflow.keras import regularizers
from tkinter import *
from tkinter import messagebox
import tkinter as tk
import matplotlib
from PythonApplication1 import matplotlib, plt, plt
import PythonApplication1
from tkinter import *
import tkinter as tk 
from tkinter import messagebox
import tkinter.filedialog as fd
from tkinter import ttk
import metods
from sklearn.preprocessing import MinMaxScaler
from tkinter import messagebox as mb
from sklearn.metrics import mean_squared_error
from keras.models import Sequential
from keras.layers import Dense, SimpleRNN, Dropout, Flatten
import time, os, argparse, io
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense
from keras.layers import Embedding
from keras.layers import LSTM, Conv1D, Dropout
from keras.layers import LeakyReLU
from keras.layers import BatchNormalization
from tensorflow import keras
from tensorflow.keras import layers
import h5py
from mpl_toolkits.mplot3d import Axes3D
from tensorflow.keras.callbacks import EarlyStopping
import openpyxl
from tensorflow.keras import Model, Input, backend
from keras.layers import TimeDistributed
from keras.callbacks import LearningRateScheduler
import os
import originpro as op
import csv
#from vtkmodules.qt.QVTKRenderWindowInteractor import QVTKRenderWindowInteractor
from PyQt5 import QtWidgets


dir = os.path.dirname(os.path.realpath(__file__))
#tf.compat.v1.disable_eager_execution()


def columUnity_x_y (x,y):
    x_0=np.ones((0))
    y_0=np.ones((0))

    for i in  range(0,(len(x)),1):
            y_0=np.append(y_0,y)
           
            for j in  range(0,(len(x)),1):
                x_0=np.append(x_0,x[i])
           
    return x_0,y_0

def columUnity (x,y,z):
    x_0=np.ones((0))
    y_0=np.ones((0))
    z_0=np.ones((0))

    for i in  range(0,(len(x)),1):
            y_0=np.append(y_0,y)
            
            z_0=np.append(z_0,z[i:,])
            for j in  range(0,(len(x)),1):
                x_0=np.append(x_0,x[i])
           
    return x_0,y_0,z_0

def columUnity_for_2 (z):
    z_0=z[:,:,0]
    print(z[1:,:,])
    print(z[:,1:,])
    print(z[:,:,1])
    sh1,sh2,sh3=np.shape(z)
    for i in  range(1,(sh1-1),1):
            
            z_0=np.append(z_0,z[i:,:,],axis = 1)
            
           
    return z_0


def mypolit_plot(model, x, y, history):


  num_rows, num_cols = x.shape
  
  if (num_cols==1):
      
      PythonApplication1.grach.draw_plot_2d (model)
   #   plt.subplot (2, 1, 1)
    #  plt.scatter(x[:,0], y)
     # xx=np.arange(min(x[:,0]), max(x[:,0]), 0.1)
     # plt.scatter(x[:,0], model.predict(x), s=1)
     # plt.subplot (2, 1, 2)
    #  plt.plot(history.history['loss'], label='train')
    #  plt.plot(history.history['val_loss'], label='test')
    #  plt.legend()
 
      #plt.show()

  if (num_cols==0):

     #min(x[:,1]), max(x[:,1]
      x_0 = np.arange(0, 15, 0.01)
      x_1 = np.arange(0, 15, 0.01)

      xd=np.ones((len(x_0),2))
      xd[:,0]=x_0
      xd[:,1]=x_1
      plt.subplot (1, 2, 1)
      plt.scatter(x[:,0], y, s=1,color='r', label = 'e')
     
     # plt.scatter(xd[:,0], model.predict(xd), s=1)
      #plt.legend()
      #plt.subplot (1, 3, 2)
      plt.scatter(x[:,1], y,s=1,color='b', label = 'fi')
     
      plt.scatter(xd[:,1], model.predict(xd), s=1,color='k')
      plt.legend() 
      
      plt.subplot (1, 2, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==3):

     #min(x[:,1]), max(x[:,1]
      x_0 = np.arange(0, 85, 0.01)
      x_1 = np.arange(0, 85, 0.01)
      x_1 = np.arange(0, 85, 0.01)
      xd=np.ones((len(x_0),3))
      xd[:,0]=x_0
      xd[:,1]=x_1
      plt.subplot (1, 2, 1)
      plt.scatter(x[:,0], y, s=1,color='r', label = 'e')
     
     # plt.scatter(xd[:,0], model.predict(xd), s=1)
      #plt.legend()
      #plt.subplot (1, 3, 2)
      plt.scatter(x[:,1], y,s=1,color='b', label = 'fi')
      plt.scatter(x[:,2], y,s=1,color='b', label = 'fi')     
      plt.scatter(xd[:,1], model.predict(xd), s=1,color='k')
      plt.legend() 
      
      plt.subplot (1, 2, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==2): #можно поставить 2, что бы рисовать поверзности
  #  if (PythonApplication1.ttt==1):
      #  fig.close()
    fig = plt.figure(figsize=(6, 5), dpi=75)
    print(x)
    print(y)

    #fig.add_subplot(111)
   # axx1 = fig.add_subplot(111)


   # ax_3d = Axes3D(fig)
   # ax_3d.set_xlabel('e')
  #  ax_3d.set_ylabel('fi')
  #  ax_3d.set_zlabel('C')
    x_0 = np.arange(min(x[:,0])-0.5, max(x[:,0])+0.5, 0.5)
    x_1 = np.arange(min(x[:,1])-0.5, max(x[:,1])+0.5, 0.005)
    
    xd=np.ones((len(x_0),2))
    xd[:,0]=x_0
   # xd[:,1]=x_1
    yd=np.ones((len(x_0),(len(x_1))))
    xshape1, yshape1=np.shape(yd)
    yd_2=np.ones(((xshape1)*(yshape1),2))
    yd_3=np.ones(((xshape1),(yshape1),2))
    r=0
    for i in range(0,(xshape1),1): 
        for j in  range(0,(yshape1),1):
            
            yd_3[i][j][0]= x_0[i]
            yd_3[i][j][1]=x_1[j]
            
            yd_2[r][0]=x_0[i]
            yd_2[r][1]= x_1[j]
            r+=1
            #yd[i][j]=model.predict(er)


   # x_1,y_1=columUnity_x_y(xd[:,0],xd[:,1])
    YYY=np.ones(((xshape1)*(yshape1),1))
    YYY=model.predict(yd_2)
    
    YYY_1=np.ones((len(x_1),len(x_0)))
    for i in  range(0,(len(x_0)),1):
          YYY_1[:,i]=YYY[(i*len(x_1)):(i*len(x_1)+len(x_1)),0]

   # ax_3d.scatter(x[:,0],x[:,1], y,s=15,color='r')
    xgrid, ygrid = np.meshgrid(x_0, x_1)

    #op.exit()

    PythonApplication1.grach.draw_plot_3d(xgrid, ygrid ,  YYY_1,x[:,0],x[:,1], y,history,model)
    #ax_3d.plot_wireframe(xgrid, ygrid ,  YYY_1)

 #   df1 = pd.DataFrame(data1,columns=['Country','GDP_Per_Capita'])

   # PythonApplication1.figure1 = plt.Figure(figsize=(8,5), dpi=75)

 #   PythonApplication1.ax1 = PythonApplication1.figure1.add_subplot(111)
   # ax_3d.plot_wireframe(xgrid, ygrid ,  YYY_1)
  #  PythonApplication1.ttt=1
  #  ax = plt.axes(projection='3d')
    #fig.add_subplot(111)
  #  ax.plot_surface(xgrid, ygrid ,  YYY_1, lw=0, cmap='copper')
   # PythonApplication1.bar2 = FigureCanvasTkAgg(fig, PythonApplication1.root)
    #PythonApplication1.bar2.get_tk_widget().pack(expand=1, anchor=SE, fill=X)


  #  df1 = df1[['Country','GDP_Per_Capita']].groupby('Country').sum()
   # df1.plot(kind='bar', legend=True, ax=ax1)
   # ax1.set_title('Country Vs. GDP Per Capita')


   # plt.show()


def origin_plot(model, x, y, history):


  num_rows, num_cols = x.shape
  if (num_cols==0):
    op.set_show()
    wks = op.new_sheet()
    wks.from_list(0, x[:,0], 'X Values')
    wks.from_list(1, y, 'Y Values')  
    gp = op.new_graph()
    gl = gp[0]
    gl.add_plot(wks, 1, 0)
    gl.rescale()

    fpath = op.path('u') + 'simple.png'
    gp.save_fig(fpath)

  if (num_cols==0):
      plt.subplot (2, 1, 1)
      plt.scatter(x[:,0], y)
      xx=np.arange(min(x[:,0]), max(x[:,0]), 0.1)
      plt.scatter(x[:,0], model.predict(x), s=1)
      plt.subplot (2, 1, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==0):

     #min(x[:,1]), max(x[:,1]
      x_0 = np.arange(0, 15, 0.01)
      x_1 = np.arange(0, 25, 0.01)

      xd=np.ones((len(x_0),2))
      xd[:,0]=x_0
      xd[:,1]=x_1
      plt.subplot (1, 2, 1)
      plt.scatter(x[:,0], y, s=1,color='r', label = 'e')
     
     # plt.scatter(xd[:,0], model.predict(xd), s=1)
      #plt.legend()
      #plt.subplot (1, 3, 2)
      plt.scatter(x[:,1], y,s=1,color='b', label = 'fi')
     
      plt.scatter(xd[:,1], model.predict(xd), s=1,color='k')
      plt.legend() 
      
      plt.subplot (1, 2, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==3):

     #min(x[:,1]), max(x[:,1]
      x_0 = np.arange(0, 85, 0.01)
      x_1 = np.arange(0, 85, 0.01)
      x_1 = np.arange(0, 85, 0.01)
      xd=np.ones((len(x_0),3))
      xd[:,0]=x_0
      xd[:,1]=x_1
      plt.subplot (1, 2, 1)
      plt.scatter(x[:,0], y, s=1,color='r', label = 'e')
     
     # plt.scatter(xd[:,0], model.predict(xd), s=1)
      #plt.legend()
      #plt.subplot (1, 3, 2)
      plt.scatter(x[:,1], y,s=1,color='b', label = 'fi')
      plt.scatter(x[:,2], y,s=1,color='b', label = 'fi')     
      plt.scatter(xd[:,1], model.predict(xd), s=1,color='k')
      plt.legend() 
      
      plt.subplot (1, 2, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==2): #можно поставить 2, что бы рисовать поверзности
   # max(x[:,0])+10
    
   
    x1=(np.arange(float(PythonApplication1.message_entry_from_x.get()),  float(PythonApplication1.message_entry_to_x.get()), float(PythonApplication1.message_entry_interval_x.get())))
    y1 =(np.arange(float(PythonApplication1.message_entry_from_y.get()),  float(PythonApplication1.message_entry_to_y.get()), float(PythonApplication1.message_entry_interval_y.get())))

    xgrid, ygrid = np.meshgrid(x1, y1)
    XYpairs = np.dstack([xgrid, ygrid]).reshape(-1, 2)
    q1,q2 = np.shape(xgrid)
       
    XYpairs2=PythonApplication1.Main_class.Model_main.predict(XYpairs)
   # XYpairs2= np.reshape(XYpairs2,(q1,q2))   
    x_0 = np.arange(0.5, 20, 0.1)
    x_1 = np.arange(0.5, 20, 0.1)

    xd=np.ones((len(x_1),2))
    xd[:,1]=x_1
  #  xd[:,1]=x_1
    yd=np.ones((len(x_0),(len(x_1))))
    xshape1, yshape1=np.shape(yd)
    yd_2=np.ones(((xshape1)*(yshape1),2))

    r=0


    for i in range(0,(xshape1),1): 
        for j in  range(0,(yshape1),1):

            yd_2[r][0]=x_0[i]
            yd_2[r][1]= x_1[j]
            r+=1
          
    xx_1,yy_1=columUnity_x_y(x_0,x_1)
    YYY=np.ones(((xshape1)*(yshape1),1))
    YYY=model.predict(yd_2)
    

    op.set_show()
    wks = op.new_sheet()
    wks.from_list(0,  XYpairs[:,0], PythonApplication1.name_colum[0])
    wks.from_list(1, XYpairs[:,1], PythonApplication1.name_colum[1])
    wks.from_list(2, XYpairs2[:,0], PythonApplication1.name_colum[2])
    wks.cols_axis('xyz')
    wks1 = op.new_sheet()
    wks1.from_list(0, x[:,0], PythonApplication1.name_colum[0])
    wks1.from_list(1, x[:,1], PythonApplication1.name_colum[1])
    wks1.from_list(2, y, PythonApplication1.name_colum[2])


    wks1.cols_axis('xyz') 

    # Plot 3D surface
    gp = op.new_graph(template='glCMAP')
    p = gp[0].add_plot(wks,coly=1,colx=0,colz=2, type=103) 
    gp[0].rescale()

 #   gl_2 = gr[1]
#    p2 = gl_2.add_plot(wks1,  type=103) # X is col A, Y is col C. 202 is Line + Symbol.
#    p2.color = '#ff5833'
 #   gl_2.rescale()
    
  #  op.exit()
    # Plot contour
  #  gp = op.new_graph(template='TriContour')
  #  p = gp[0].add_plot(wks,coly=1,colx=0,colz=2, type=243)
   # p.colormap = 'Maple.pal'
   
def CSV_plot(model, x, y,history):
    
  num_rows, num_cols = x.shape
  if (num_cols==2):
    x_0 = np.arange(min(x[:,0])-0.5, max(x[:,0])+0.5, 0.5)
    x_1 = np.arange(min(x[:,1])-0.5, max(x[:,1])+0.5, 0.5)

    xd=np.ones((len(x_0),2))
    xd[:,0]=x_0
 #   xd[:,1]=x_1
    yd=np.ones((len(x_0),(len(x_1))))
    yd_1=np.ones((len(x_0),(len(x_1)),2))
    xshape1, yshape1=np.shape(yd)
    r=0
    yd_2=np.ones(((xshape1)*(yshape1),2))

    for i in range(0,(len(x_0)),1): 
        for j in  range(0,(len(x_1)),1): 

            yd_2[r][0]=x_0[i]
            yd_2[r][1]= x_1[j]
            r+=1
          
    x_11,y_1=columUnity_x_y(x_0,x_1)
    YYY=np.ones(((xshape1)*(yshape1),1))
    YYY=model.predict(yd_2)
   # x_1,y_1,z_1 = columUnity (xd[:,0],xd[:,1],yd)
    с_1=np.ones((len(YYY[:,0]),3))
    
    #c = np.concatenate((x_1, y_1), axis=1)
    с_1[:,0] =yd_2[:,0]
    с_1[:,1] =yd_2[:,1]
    с_1[:,2] =YYY[:,0]
    #df = pd.from_csv('E:\employee_file.csv')
    cities = pd.DataFrame(с_1, columns=['X', 'Y', 'Z'])
    os.remove('E:\employee_file.csv')
    cities.to_csv('E:\employee_file.csv', index=False)

        




def network_plot (model, x, y,history):
  if PythonApplication1.comboExample1.get() == "Matplotlib":
             mypolit_plot(model, x, y,history)
  if PythonApplication1.comboExample1.get() == "Ориджин":
             origin_plot(model, x, y,history)
  if PythonApplication1.comboExample1.get() == "CSV":
             CSV_plot(model, x, y,history)




  num_rows, num_cols = x.shape
  if (num_cols==0):
    op.set_show()
    wks = op.new_sheet()
    wks.from_list(0, x[:,0], 'X Values')
    wks.from_list(1, y, 'Y Values')  
    gp = op.new_graph()
    gl = gp[0]
    gl.add_plot(wks, 1, 0)
    gl.rescale()

    fpath = op.path('u') + 'simple.png'
    gp.save_fig(fpath)

  if (num_cols==0):
      plt.subplot (2, 1, 1)
      plt.scatter(x[:,0], y)
      xx=np.arange(min(x[:,0]), max(x[:,0]), 0.1)
      plt.scatter(x[:,0], model.predict(x), s=1)
      plt.subplot (2, 1, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==0):

     #min(x[:,1]), max(x[:,1]
      x_0 = np.arange(0, 15, 0.01)
      x_1 = np.arange(0, 15, 0.01)

      xd=np.ones((len(x_0),2))
      xd[:,0]=x_0
      xd[:,1]=x_1
      plt.subplot (1, 2, 1)
      plt.scatter(x[:,0], y, s=1,color='r', label = 'e')
     
     # plt.scatter(xd[:,0], model.predict(xd), s=1)
      #plt.legend()
      #plt.subplot (1, 3, 2)
      plt.scatter(x[:,1], y,s=1,color='b', label = 'fi')
     
      plt.scatter(xd[:,1], model.predict(xd), s=1,color='k')
      plt.legend() 
      
      plt.subplot (1, 2, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==0):

     #min(x[:,1]), max(x[:,1]
      x_0 = np.arange(0, 85, 0.01)
      x_1 = np.arange(0, 85, 0.01)
      x_1 = np.arange(0, 85, 0.01)
      xd=np.ones((len(x_0),3))
      xd[:,0]=x_0
      xd[:,1]=x_1
      plt.subplot (1, 2, 1)
      plt.scatter(x[:,0], y, s=1,color='r', label = 'e')
     
     # plt.scatter(xd[:,0], model.predict(xd), s=1)
      #plt.legend()
      #plt.subplot (1, 3, 2)
      plt.scatter(x[:,1], y,s=1,color='b', label = 'fi')
      plt.scatter(x[:,2], y,s=1,color='b', label = 'fi')     
      plt.scatter(xd[:,1], model.predict(xd), s=1,color='k')
      plt.legend() 
      
      plt.subplot (1, 2, 2)
      plt.plot(history.history['loss'], label='train')
      plt.plot(history.history['val_loss'], label='test')
      plt.legend()
 
      plt.show()

  if (num_cols==0): #можно поставить 2, что бы рисовать поверзности
    #fig = plt.figure(figsize=(7, 4))
   # ax_3d = Axes3D(fig)
   # ax_3d.set_xlabel('fi')
   # ax_3d.set_ylabel('e')
   # ax_3d.set_zlabel('C')
    x_0 = np.arange(1.5, 30, 0.1)
    x_1 = np.arange(1.5, 30, 0.1)

    xd=np.ones((len(x_0),2))
    xd[:,0]=x_0
    xd[:,1]=x_1
    yd=np.ones((len(x_0),(len(x_0))))
    for i in range(0,(len(x_0)-1),1): 
        for j in  range(0,(len(x_0)-1),1):
            er=np.ones((1,2))
            er[0][0]= xd[j,0]
            er[0][1]= xd[i,1]
            yd[i][j]=model.predict(er)
    
    #ax_3d.scatter(x[:,0],x[:,1], y,s=7,color='r')
    #xgrid, ygrid = np.meshgrid(xd[:,0], xd[:,1])
    x_1,y_1,z_1 = columUnity (xd[:,0],xd[:,1],yd)
    op.set_show()
    wks = op.new_sheet()
    wks.from_list(0, x_1, 'fi Values')
    wks.from_list(1, y_1, 'e Values')
    wks.from_list(2, z_1, 'C Values')
    wks.cols_axis('xyz') 

    # Plot 3D surface
    gp = op.new_graph(template='glCMAP')
    p = gp[0].add_plot(wks,coly=1,colx=0,colz=2, type=103) 
    gp[0].rescale()

    # Plot contour
    gp = op.new_graph(template='TriContour')
    p = gp[0].add_plot(wks,coly=1,colx=0,colz=2, type=243)
    p.colormap = 'Maple.pal'
    


    #op.exit()


    #ax_3d.plot_wireframe(xgrid, ygrid,  yd)

    #plt.show()


def mix_oint(x_files,y_files):  #Перетасовка точек

 x_rows, x_cols = x_files.shape
 x_y_shuffle=np.column_stack([x_files, y_files]) 
 num_rows, num_cols = x_y_shuffle.shape
 #x_y_shuffle=x_y_shuffle.T
 np.random.shuffle(x_y_shuffle)
 #x_y_shuffle=x_y_shuffle.T
 y=x_y_shuffle[:,num_cols-1]
 x = np.delete(x_y_shuffle, np.s_[-1:], axis=1)
 return x,y #перемешивание точек


def fit_model(model,x, y):
 #x1,y1=mix_oint(x,y)
# es = EarlyStopping(monitor='loss', mode='min', verbose=0, patience=300)
 es = EarlyStopping(monitor='val_loss', mode='min', verbose=0, patience=50000)
 mc = keras.callbacks.ModelCheckpoint('best_model.h5', monitor='val_acc', mode='max', verbose=0, save_best_only=True)




 history = model.fit(x, y, epochs=90000, batch_size=400, validation_split=0.15,validation_freq=2, callbacks=[es, mc])
 return history, model #обучение модели


def file_acceptance (): #считывание данных из файла


 url=PythonApplication1.message.get()
 if (url.find(".xlsx",len(url)-5) !=-1):
    WS = pd.read_excel(url)
    WS_np = np.array(WS)
    print(WS.columns.ravel())
    num_rows, num_cols = WS_np.shape
    PythonApplication1.name_colum=WS.columns.ravel()
    x_files=np.ones((num_rows,num_cols-1))

   #  for number in range(1,num_cols-1,1):
     #    x[:,number]
    #    np.append(x, datnp[number], axis=1)
    y_files=WS_np[:,num_cols-1]
    x_files = np.delete(WS_np, np.s_[-1:], axis=1)
    return x_files, y_files
    #wb_obj = openpyxl.load_workbook(filename = url)

    #sheet_obj = wb_obj.active #Выбираем активный лист таблицы(
    #m_row = sheet_obj.max_row

    #for i in range(1, m_row + 1):
    #    cell_obj1 = sheet_obj.cell(row=i, column=1) # В column= подставляем номер нужной колонки
    #    cell_obj2 = sheet_obj.cell(row=i, column=2)
    
     #   x.append(cell_obj1.value)
    #    y.append(cell_obj2.value)
 if (url.find(".csv",len(url)-5) !=-1):
     dff = pd.read_csv(url, decimal='.', delimiter=',', dayfirst=True)
    
     datnp=dff.to_numpy()
     num_rows, num_cols = datnp.shape
     x_files=np.ones((len(datnp[:,0]),num_cols-1))

   #  for number in range(1,num_cols-1,1):
     #    x[:,number]
    #    np.append(x, datnp[number], axis=1)
     y_files=datnp[:,num_cols-1]
     x_files = np.delete(datnp, np.s_[-1:], axis=1)
     return x_files, y_files
   #  x=np.delete(datnp,1, num_cols-1)



 #return x_files,y_files

 #dff = pd.read_table(url, names=['val','vale'],decimal='.', delimiter=',')


def postroenie ():#��������
 #url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vQ_Mv5MB1Ty1ixEnsXeDJwz0_31rQ1woLrfvmT-SSRuqcUybp6cTy8CgqncjtYxC41ZG5HlzPqUN0au/pub?gid=0&single=true&output=csv'
 
 url=PythonApplication1.message.get()
 df = pd.read_csv(url, names=['val','vale'],decimal='.', delimiter=',', dayfirst=True)



 #xxx=df['val'].to_numpy()
 #yyy=df['vale'].to_numpy()

 #xx = np.array([df['val'].to_numpy()], dtype=float)
 #yy = np.array([df['vale'].to_numpy()], dtype=float)
 # Define the number of nodes
 n_nodes_hl1 = 100
 n_nodes_hl2 = 100


 # Define the number of outputs and the learn rate
 n_classes = 1
 

 #xx.dtype
 learning_rate = 0.01
 training_epochs = 100

 #x_train=np.array(df['val'].to_numpy(), dtype=float)
 #y_train=np.array(df['vale'].to_numpy(), dtype=float)

 x_train = np.linspace(-1, 1, 101)
 y_train = 2 * x_train + np.random.randn(*x_train.shape) * 0.33

 X = tf.placeholder(tf.float32)
 Y = tf.placeholder(tf.float32)
 
 def model(X, w):
    return tf.multiply(X, w) 
 
    
 w = tf.Variable(0.0, name="weights")
 
 y_model = model(X, w)
 cost = tf.square(Y-y_model)
 
 train_op = tf.train.GradientDescentOptimizer(learning_rate).minimize(cost)
 
 sess = tf.Session()
 init = tf.global_variables_initializer()
 sess.run(init)
    
 for epoch in range(training_epochs):
   for (x, y) in zip(x_train, y_train):
     sess.run(train_op, feed_dict={X: x, Y: y})
 
 w_val = sess.run(w)
 
 sess.close()
 plt.scatter(x_train, y_train)
 y_learned = x_train*w_val
 plt.plot(x_train, y_learned, 'r')
 plt.show()

def postroeniepolin (): #������� ������ �� ��������
 #url = 'https://docs.google.com/spreadsheets/d/e/2PACX-1vQ_Mv5MB1Ty1ixEnsXeDJwz0_31rQ1woLrfvmT-SSRuqcUybp6cTy8CgqncjtYxC41ZG5HlzPqUN0au/pub?gid=0&single=true&output=csv'
 tf.disable_v2_behavior()  
 url=PythonApplication1.message.get()
 df = pd.read_csv(url, names=['val','vale'],decimal='.', delimiter=',', dayfirst=True)
 num_coeffs=6
 trY_coeffs= [1, 2, 3, 4, 5, 6]
 learning_rate = 0.01
 training_epochs = 40
 trY=0

 #xxx=df['val'].to_numpy()
 #yyy=df['vale'].to_numpy()

 #trX = np.array([df['val'].to_numpy()], dtype=float)
 #trY = np.array([df['vale'].to_numpy()], dtype=float)
 
 """
 trX = np.linspace(-1, 1, 101)

 for i in range(num_coeffs):
    
    trY += trY_coeffs[i] * np.power(trX, i)
 trY += np.random.randn(*trX.shape) * 1.5
 """
 

 #trX=df['val'].to_numpy()
 #trY=df['vale'].to_numpy()

 trX = np.array(df['val'].to_numpy(), dtype=float)
 trY = np.array(df['vale'].to_numpy(), dtype=float)

 plt.scatter(trX, trY)


 X = tf.placeholder(tf.float32)
 Y = tf.placeholder(tf.float32)

 def model(X, w):
    terms = []
    for i in range(num_coeffs):
        term = tf.multiply(w[i], tf.pow(X, i))
        terms.append(term)
    return tf.add_n(terms)

 w = tf.Variable([0.] * num_coeffs, name="parameters")
 y_model = model(X, w)
  
 cost = (tf.pow(Y-y_model, 2))
 train_op = tf.train.GradientDescentOptimizer(learning_rate).minimize(cost)
 
 sess = tf.Session()
 init = tf.global_variables_initializer()
 sess.run(init)
 
 for epoch in range(training_epochs):
    for (x, y) in zip(trX, trY):
        sess.run(train_op, feed_dict={X: x, Y: y})
 
 w_val = sess.run(w)
 print(w_val)
 sess.close()
 plt.scatter(trX, trY)
 trY2 = 0
 for i in range(num_coeffs):
    trY2 += w_val[i] * np.power(trX, i)


 plt.plot(trX, trY2, 'r')
 plt.show()
 
def probanerset ():
 #if __name__ == '__main__':
    predictor = SeriesPredictor(input_dim=1, seq_size=4, hidden_dim=10)
    train_x = [[[1], [2], [5], [6]],
               [[5], [7], [7], [8]],
               [[3], [4], [5], [7]]]
    train_y = [[1, 3, 7, 11],
               [5, 12, 14, 15],
               [3, 7, 9, 12]]
    predictor.train(train_x, train_y)
    test_x = [[[1], [2], [3], [4]],
              [[4], [5], [6], [7]]]
    predictor.test(test_x)

def NewOneNetwork (): 


 x,y=file_acceptance()



 model = Sequential()
 x_rows, x_cols = x.shape
 model.add(Dense(x_cols*20, activation='elu')) 
 #model.add(Dense(x_cols*40, activation='linear'))
 model.add(Dense(x_cols*20, activation='relu'))
 #model.add(Dense(x_cols*40, activation='tanh'))

 model.add(Dense(1,'elu'))


 model.compile(loss='mean_squared_error', optimizer='RMSprop')
 history, model=fit_model(model,x, y)
 
 network_plot (model, x, y,history)









def NewFourNetwork(): #сеть LSTM
 url='E:/data3.csv'
 #url=PythonApplication1.message.get()
 dff = pd.read_csv(url, names=['val','vale'],decimal='.', delimiter=',', dayfirst=True)
 #x_data = np.linspace (-0.5, 0.5, 200) [:, np.newaxis] # ������������ 200 �����, ���������� �������������� �� -0,5 �� 0,5, �������� ������� �� 200 ����� � ������ �������
 #noise = np.random.normal (0, 0.02, x_data.shape) # ��������� ���������� ����
 #y_data = np.square(x_data) + noise  # y = x^2 + noise

 #x_data=df['val'].to_numpy()
 #y_data=df['vale'].to_numpy()

 #x_data = np.array([df['val'].to_numpy()], dtype=float)
 #y_data = np.array([df['vale'].to_numpy()], dtype=float)
    
    
 #x = tf.placeholder (tf.float32, [None, 1]) # ���������� �����������
 #y = tf.placeholder(tf.float32, [None, 1])


 #x = -50 + np.random.random((25000,1))*100
# y = x**2
 data_dim = 16

 timesteps = 8

 num_classes = 10

 batch_size = 32
 
 x=dff['val'].to_numpy()
 y=dff['vale'].to_numpy()
# Define model
 
 few_neurons=len(x)
 model = Sequential()
 model.add(Dense(few_neurons,input_shape=(1,), input_dim=1, activation='relu'))
 #model.add(Dense(few_neurons, activation='relu'))
 model.add(Dense(few_neurons*2))

 #model.add(Dense(10, activation='relu'))
 model.add(Dense(1))
 model.compile(loss='mean_squared_error', optimizer='adam')
 model.fit(x, y, epochs=2700, batch_size=2000)





 plt.scatter(x[:,0], y)
 #xx=[]


 xx=np.arange(min(x[:,0]), (max(x[:,0])+max(x[:,0])//2), 0.1)
 plt.scatter(xx, model.predict(xx), s=1)
 plt.show() 



def make_model(input_shape):
    input_layer = keras.layers.Input(input_shape)

    conv1 = keras.layers.Conv1D(filters=64, kernel_size=3, padding="same")(input_layer)
    conv1 = keras.layers.BatchNormalization()(conv1)
    conv1 = keras.layers.ReLU()(conv1)

    conv2 = keras.layers.Conv1D(filters=64, kernel_size=3, padding="same")(conv1)
    conv2 = keras.layers.BatchNormalization()(conv2)
    conv2 = keras.layers.ReLU()(conv2)

    conv3 = keras.layers.Conv1D(filters=64, kernel_size=3, padding="same")(conv2)
    conv3 = keras.layers.BatchNormalization()(conv3)
    conv3 = keras.layers.ReLU()(conv3)

    gap = keras.layers.GlobalAveragePooling1D()(conv3)

    output_layer = keras.layers.Dense(num_classes, activation="softmax")(gap)

    return keras.models.Model(inputs=input_layer, outputs=output_layer)


def step_decay(epoch):
    lrate = 0.001
    if epoch > 16:
        lrate = 0.0005
    if epoch > 32:
        lrate = 0.0001
    if epoch > 45:
        lrate = 0.00005
    return lrate


def split_sequence(sequence, n_steps):
	X, y = list(), list()
	for i in range(len(sequence)):
		# find the end of this pattern
		end_ix = i + n_steps
		# check if we are beyond the sequence
		if end_ix > len(sequence)-1:
			break
		# gather input and output parts of the pattern
		seq_x, seq_y = sequence[i:end_ix], sequence[end_ix]
		X.append(seq_x)
		y.append(seq_y)
	return np.array(X), np.array(y)



def RNN (): #ФУНКЦИЯ,

 #url='E:/data3.csv'
 x,y=file_acceptance()



 x_y_shuffle=np.ones((2,len(x_files))) #Перетасовка точек

 x_y_shuffle[0]=x_files.copy()

 x_y_shuffle[1]=y_files.copy()

 x_y_shuffle=x_y_shuffle.T

 #np.random.shuffle(x_y_shuffle)

 x_y_shuffle=x_y_shuffle.T

 x=x_y_shuffle[0]

 y=x_y_shuffle[1]

 x_training1=np.split(x, [0, (len(x)//4)*3])#разделение массивов на обучающую и тестовые выборки ( в пропорции75%/25%)

 y_training1=np.split(y, [0, (len(x)//4)*3])# с помощь команды split можно разделить массив с n-го элемента до m-го элемента

 x=x_training1[1]

 y=y_training1[1]

 x_test=x_training1[2]

 y_test=y_training1[2]

 look_back=1
 step = 4
 top_words = 5000
 max_review_length = 500
 embedding_vecor_length = 32
# lx=len(x)
 x = np.expand_dims(x, 1) 
# y = np.expand_dims(y, 1)

 y_raz,y1_raz = split_sequence(y, 3)


 model = Sequential()


 model.add(layers.Embedding(input_dim=1024, output_dim=4))
 model.add(layers.GRU(128, return_sequences=True,activation="tanh",dropout=0.5))
 model.add(LSTM(64,activation="tanh",dropout=0.1 , input_shape=(1, 1)))
    #model.add(SimpleRNN(256))
# model.add(Dropout(0.2))
# model.add(SimpleRNN(256,activation="tanh",dropout=0.8))

# model.add(Dropout(0.2))
 model.add(Dense(100, activation='tanh'))
 model.add(Dense(1, activation='elu'))

 model.compile(loss='mse', optimizer='adam',metrics=['accuracy'])
 model.summary()
 reduce_lr = keras.callbacks.ReduceLROnPlateau(monitor='val_loss', factor=0.2,
                              patience=5, min_lr=0.001)
 model.fit(x, y, epochs=500, batch_size=50,verbose=2, 
          validation_data=(x_test, y_test),
          shuffle=True,
          callbacks=[reduce_lr],validation_split=0.1)


 plt.scatter(x, y, s=7)
 plt.scatter(x_test, y_test, s=7)
 xx=np.arange(min(x_files), (max(x_files)+max(x_files)//2), 0.01)

 plt.scatter(xx, model.predict(xx), s=1)
 plt.show() 


def stable_network():
 model = Sequential()
 initializer = keras.initializers.HeNormal(seed=None)
 model.add(Dense(1, input_dim=1, activation='tanh'))
 model.add(Dense(1024, activation='tanh'))
 model.add(Dense(512, activation='tanh')) 
 model.add(Dense(1, activation='elu'))

 sgd = keras.optimizers.SGD(lr=0.01, momentum=0.9, nesterov=True)
 model.compile(loss='mean_squared_error', optimizer='adam')
 return model



def make_model_3(): 
    model = Sequential()
    initializer = keras.initializers.HeNormal(seed=None)
    model.add(Dense(512,  activation='tanh'))
    model.add(Dense(1024, activation='tanh'))
    model.add(Dense(512, activation='linear')) 
    model.add(Dense(1,'relu'))
    model.compile(loss='mean_squared_error', optimizer='adam')
    return model

def make_model_4(trained_model): #ОШИБКА РАЗМЕРНОСТИ
    model = Sequential()
    q51=np.zeros((1,1))
    q51[0][0]=trained_model.get_weights()[7]
    initializer = keras.initializers.HeNormal(seed=None)
    den=np.ones((1,1))
    dene=den.tolist()


    listOfNumpyArrays = [np.empty(shape = (1,1), dtype = np.float32)]
    listOfNumpyArrays[0][0] = trained_model.get_weights()[7]




    model.add(Dense(1, input_dim=1, activation='relu', use_bias=False, trainable=False, weights=listOfNumpyArrays))

    print('1 model:')
    print(listOfNumpyArrays[0][0])
    print('2 model:')
    print(model.get_weights()[0])
    q52=q51[0:]

    model.add(Dense(1024, activation='tanh'))
    model.add(Dense(512, activation='tanh')) 
    model.add(Dense(1, activation='relu'))
    model.compile(loss='mean_squared_error', optimizer='adam')
    return model

    

def make_model_2(trained_model):
    
    inp = Input(shape=(1,), name='dense_1')
    #inp.set_weights(trained_model.get_weights()[4])
    m = make_model_1()
   # m.set_weights(trained_model.get_weights())
    #print (np.shape(trained_model.get_weights()[5]))
    q51=(trained_model.get_weights()[4].copy)
 #   inp.set_weights(trained_model.get_weights()[0])
    l2 = Dense(300, activation = 'relu',weights=[q51], trainable=False)(inp)

    l3 = Dense(600, activation = 'relu')(l2)
   

    out = Dense(1, activation = 'elu')(l3)
    #bucket = tf.stack([out1, out2], axis=2)
    #out = tf.keras.backend.squeeze(Dense(1)(bucket), axis = 2)
    model2 = Model(inp, out)
    q5=(trained_model.get_weights()[4])
    
    q66=(model2.get_weights()[0])
    print(model2.get_weights()[0])
 
    q77=q5[1]
    
    #model2.layers[0].set_weights([q5.T, np.ones(model2.layers[0].get_weights()[0].shape)])
    return model2

def make_model_1():
    
    inp = Input(1)

    l1 = Dense(200, activation = 'sigmoid')(inp)


    l3 = Dense(300, activation = 'sigmoid')(l1)
    out3 = Dense(1)(l3)
    model1 = Model(inp, out3)
    
    return model1

def TandemNN ():
 x,y=file_acceptance()
 
 model1 = make_model_3()
 model1.compile(optimizer = keras.optimizers.Adam(),
               loss = keras.losses.mean_squared_error)
 fit_model(model1,x, y)

 model2 = make_model_4(model1)
 model2.summary()

 model2.compile(optimizer = keras.optimizers.Adam(),
               loss = keras.losses.mean_squared_error)

 fit_model(model2,x, y)

 network_plot (model2, x, y)
 

def conclusion_MSE(a): #выводит сообщение
        msg = a
        mb.showinfo("MSE", msg)

def PodborZnacheny():# тандем НС

 #url='E:/data3.csv'
 x,y=file_acceptance()

 fun_active_all=['relu','tanh','elu','softmax','selu','softplus','softsign','sigmoid','hard_sigmoid','linear'] 
 fun_optimizers_all=['RMSprop', 'sgd', 'adam', 'Nadam','Adamax','Adadelta', 'Adagrad']
 x_files=dff['val'].to_numpy()
 y_files=dff['vale'].to_numpy()
# Define model
 
 x=np.ones(len(x_files))
 y=np.ones(len(y_files))#Создание массивов

 x_y_shuffle=np.ones((2,len(x_files))) #Перетосовка точек 
 x_y_shuffle[0]=x_files.copy()
 x_y_shuffle[1]=y_files.copy()
 x_y_shuffle=x_y_shuffle.T
 np.random.shuffle(x_y_shuffle)
 x_y_shuffle=x_y_shuffle.T
 x=x_y_shuffle[0]
 y=x_y_shuffle[1]

 temp = list(zip(x_y_shuffle[0], x_y_shuffle[1]))


 separation_array=(len(x)//4)*3


 x_training1=np.split(x, [0, separation_array])
 y_training1=np.split(y, [0, separation_array])

 x_training=x_training1[1]
 y_training=y_training1[1]

 x_test=x_training1[2]
 y_test=y_training1[2]
 i=0
 min_index=1000
 min_element=1000
 search_min_mse=np.ones([101])
 while i<=100:
  search_min_mse[i]=1000
  i+=1

  
     
     
 i=0
 while i<=len(fun_active_all)-1:




  few_neurons=len(x)
  model = Sequential()
 # tens_i=i

  model.add(Dense(512,input_shape=(1,), input_dim=1, activation=fun_active_all[i]))
  model.add(Dense(1024, activation=fun_active_all[i]))
 # model.add(Dense(1024, activation=fun_active_all[i]))
  model.add(Dense(512, activation=fun_active_all[i]))
  model.add(Dense(1))

 #model.add(Dense(1))
 
 
 #model.add(Dense(200, activation='sigmoid'))
 #model.add(Dense(200, activation='sigmoid'))
 #model.add(Dense(100, activation='sigmoid'))
 #model.add(Dense(5, activation='sigmoid'))
 #model.add(Dense(10, activation='relu'))

  
  model.compile(loss='mean_squared_error', optimizer= 'RMSprop')

  model.fit(x_training, y_training, epochs=200+i, batch_size=512)

   #search_min_mse[i]=mean_squared_error(y_test, model.predict(x_test))
  if (mean_squared_error(y_test, model.predict(x_test))<min_element):
     min_element=mean_squared_error(y_test, model.predict(x_test))
     min_index=i
    
  i +=1
  

 mse_a=min_index

 model = Sequential()

 model.add(Dense(512,input_shape=(1,), input_dim=1, activation=fun_active_all[mse_a]))
 model.add(Dense(1024, activation=fun_active_all[mse_a]))
# model.add(Dense(1024, activation=fun_active_all[mse_a]))
 model.add(Dense(512, activation=fun_active_all[mse_a]))


 model.add(Dense(1))
 model.compile(loss='mean_squared_error', optimizer= 'RMSprop')

 model.fit(x_training, y_training, epochs=200, batch_size=200)

 plt.scatter(x_training, y_training)
 #xx=[]
 plt.scatter(x_test, y_test)
 
 conclusion_MSE(fun_active_all[min_index])

 xx=np.arange(min(x), (max(x)+max(x)//2), 0.1)

 plt.scatter(xx, model.predict(xx), s=1)
 plt.show() 


def choiceCombobox():#выбор функции
    # PythonApplication1.figure1.clf()
     if PythonApplication1.comboExample.get() == "probanerset":
             probanerset()
     if PythonApplication1.comboExample.get() == "Lineq":
             postroenie()
     if PythonApplication1.comboExample.get() == "Polinel":
             postroeniepolin()
     if PythonApplication1.comboExample.get() == "NewOneNetwork":
             NewOneNetwork()
     if PythonApplication1.comboExample.get() == "RNN":
             RNN()
     if PythonApplication1.comboExample.get() == "TandemNN":
             TandemNN()
     if PythonApplication1.comboExample.get() == "NewFourNetwork":
             NewFourNetwork()
     if PythonApplication1.comboExample.get() == "NewFiveNetwork":
             NewFiveNetwork()
     if PythonApplication1.comboExample.get() == "PodborZnacheny":
             PodborZnacheny()